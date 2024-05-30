# type: ignore

import importlib
import typer
from typing_extensions import Annotated
from tinydb import TinyDB, Query, where
from tinydb.table import Document
from rich import print
import os
from os import path, listdir
import json
from typing import Dict
from tests.BaseTests import MessageTest, GameTest, ManualMessageTest, ManualIndependentMessageTest, ManualDependentMessageTest, TestType
from openpyxl import Workbook
from openpyxl.styles import Font

from DBTypes import Game, GameType, Message, Model, Test, TestResult

app = typer.Typer()
db = TinyDB('db.json')
game_types_table = db.table('game_types')
games_table = db.table('games')
models_table = db.table('models')
messages_table = db.table('messages')
tests_table = db.table('tests')
test_results_table = db.table('test_results')


def load_tests() -> Dict[str, Dict[str, TestType]]:
    test_groups: Dict[str, Dict[str, TestType]] = {}

    test_files = listdir("tests")

    for test_group in test_files:
        absolute_path = path.join(path.abspath("tests"), test_group)
        if not path.isfile(absolute_path) or not test_group.endswith(".py") or test_group == "BaseTests.py":
            continue

        module_name = test_group[:-3]

        test_groups[module_name] = {}
        module = importlib.import_module(f"tests.{module_name}")
        for name in dir(module):
            obj = getattr(module, name)
            if isinstance(obj, type) and issubclass(obj, (GameTest, MessageTest)) and obj != GameTest and obj != MessageTest and obj != ManualMessageTest and obj != ManualDependentMessageTest and obj != ManualIndependentMessageTest:
                test_name = f"{module_name}.{obj.__name__}"
                test_groups[module_name][test_name] = obj()

                tests_docs = tests_table.search(where('name') == test_name)
                if len(tests_docs) == 0:
                    test = Test(test_name)
                    test.insert_into_table(tests_table)

    return test_groups


test_groups = load_tests()


@app.command()
def load_chatlogs():
    print("Loading chatlogs")
    logfiles = get_logfiles()
    for logfile in logfiles:
        load_logfile(logfile)
        new_path = path.join(path.abspath(
            "chatlogs/loaded"), path.basename(logfile))
        os.rename(logfile, new_path)

    print("Chatlogs loaded")


@app.command()
def run_tests(test_group_name: Annotated[str, typer.Option(prompt=True)], game_name: Annotated[str, typer.Option(prompt=True)], run_manual_tests: Annotated[bool, typer.Option(prompt=True)] = False):
    if test_group_name not in test_groups:
        print(f"No tests found named {test_group_name}")
        return

    test_group = test_groups[test_group_name]

    game_type = game_types_table.search(where('name') == game_name)

    if len(game_type) == 0:
        print(f"No game found named {game_name}")
        return

    game_type = GameType.from_dict(game_type[0])

    run_tests_on_game_type(game_type, test_group, run_manual_tests)


@app.command()
def export():
    print("Exporting database")
    wb = Workbook()

    bold_font = Font(bold=True)

    game_types_sheet_name = "Game Types"
    models_sheet_name = "Models"
    tests_sheet_name = "Tests"
    games_sheet_name = "Games"
    test_results_sheet_name = "Test Results"
    messages_sheet_name = "Messages"

    wb.create_sheet(game_types_sheet_name)

    game_types_sheet = wb[game_types_sheet_name]

    game_types_sheet.append(["ID", "Name"])

    for cell in game_types_sheet[1]:
        cell.font = bold_font

    for game_type in game_types_table.all():
        game_types_sheet.append([game_type.doc_id, game_type['name']])

    wb.create_sheet(models_sheet_name)

    models_sheet = wb[models_sheet_name]

    models_sheet.append(["ID", "Name"])

    for cell in models_sheet[1]:
        cell.font = bold_font

    for model in models_table.all():
        models_sheet.append([model.doc_id, model['name']])

    wb.create_sheet(tests_sheet_name)

    tests_sheet = wb[tests_sheet_name]

    tests_sheet.append(["ID", "Name"])

    for cell in tests_sheet[1]:
        cell.font = bold_font

    for test in tests_table.all():
        tests_sheet.append([test.doc_id, test['name']])

    wb.create_sheet(games_sheet_name)

    games_sheet = wb[games_sheet_name]

    games_sheet.append(
        ["ID", "Game Type", "Game Type Name", "Players", "Player Names"])

    for cell in games_sheet[1]:
        cell.font = bold_font

    for i, game in enumerate(games_table.all()):
        row = i + 2
        games_sheet.append([game.doc_id, game['game_type'], f"=VLOOKUP(B{row};'{game_types_sheet_name}'!$A$2:$B${len(game_types_table) + 1};2)", ",".join(
            [str(player) for player in game['players']]), f"=JOIN(\", \";(ARRAYFORMULA(VLOOKUP(SPLIT(D{row};\",\");'{models_sheet_name}'!$A$2:$B${len(models_table) + 1};2))))"])

    wb.create_sheet(test_results_sheet_name)

    test_results_sheet = wb[test_results_sheet_name]

    test_results_sheet.append(["ID", "Test", "Test Name", "Message", "Message Content", "Result"])

    for cell in test_results_sheet[1]:
        cell.font = bold_font

    for i, test_result in enumerate(test_results_table.all()):
        row = i + 2

        test_results_sheet.append([test_result.doc_id, test_result['test'], f"=VLOOKUP(B{row};{tests_sheet_name}!$A$2:$B${len(tests_table) + 1};2)",
                                  test_result['message'], f"=VLOOKUP(D{row};'{messages_sheet_name}'!$A$2:$J${len(messages_table) + 1};6)",test_result['result']])

    wb.create_sheet(messages_sheet_name)

    messages_sheet = wb[messages_sheet_name]

    messages_sheet.append(["ID", "Sender", "Sender Name",
                          "Game", "Game Type", "Content", "Turn", "Passed Tests", "Failed Tests", "Percentage"])

    for cell in messages_sheet[1]:
        cell.font = bold_font

    for i, message in enumerate(messages_table.all()):
        row = i + 2

        messages_sheet.append([message.doc_id, message['sender'], f"=VLOOKUP(B{row};{models_sheet_name}!$A$2:$B${len(
            models_table) + 1};2)", message['game'], f"=VLOOKUP(D{row};{games_sheet_name}!$A$2:$F${len(games_table) + 1};3)", message['content'], message['turn'], f"=COUNTIFS('{test_results_sheet_name}'!$D$2:$D${len(test_results_table) + 1};A{row};'{test_results_sheet_name}'!$F$2:$F${len(test_results_table) + 1};TRUE)", f"=COUNTIFS('{test_results_sheet_name}'!$D$2:$D${len(test_results_table) + 1};A{row};'{test_results_sheet_name}'!$F$2:$F${len(test_results_table) + 1};FALSE)", f"=H{row}/(H{row}+I{row})"])

    wb.save("export.xlsx")
    print("Database exported")


def get_logfiles() -> list[str]:
    logfiles = []

    for file in listdir("chatlogs"):
        absolute_path = path.join(path.abspath("chatlogs"), file)
        if path.isfile(absolute_path) and file.endswith(".json"):
            logfiles.append(absolute_path)
    return logfiles


def load_logfile(logfile: str):
    print(f"Loading {path.basename(logfile)} into database")
    with open(logfile, "r") as f:
        data = json.load(f)
        game_name = path.basename(f.name).split(".")[0]
        game_type_docs = game_types_table.search(where('name') == game_name)
        game_type = None

        if len(game_type_docs) == 0:
            game_type = GameType(game_name)
            game_type.insert_into_table(game_types_table)
        else:
            game_type = GameType.from_dict(game_type_docs[0])

        for game in data:
            load_game(game_type.get_id(), game)


def load_game(game_type: int, game_data: Dict):
    player_to_id: Dict[str, int] = {}
    for player in game_data['players']:
        model_docs = models_table.search(where('name') == player['model'])
        model = None

        if len(model_docs) == 0:
            model = Model(player['model'])
            model.insert_into_table(models_table)
        else:
            model = Model.from_dict(model_docs[0])

        player_to_id[player['name']] = model.get_id()

    players = [player_to_id[player['name']] for player in game_data['players']]
    game = Game(game_type, players)
    game.insert_into_table(games_table)

    for message in game_data['messages']:
        sender = player_to_id[message['sender']]

        message = Message(sender, game.get_id(),
                          message['content'], message['turn'])
        message.insert_into_table(messages_table)


def run_tests_on_game_type(game_type: GameType, tests_dict: Dict[str, TestType], run_manual_tests: bool):
    games = [Game.from_dict(game_doc) for game_doc in games_table.search(
        where('game_type') == game_type.get_id())]
    tests = [Test.from_dict(test_doc) for test_doc in tests_table.search(
        where('name').one_of(list(tests_dict.keys())))]

    for test in tests:
        if not issubclass(tests_dict[test.name].__class__, ManualMessageTest):
            for game in games:
                run_test_on_game(game, test, tests_dict[test.name])

    if run_manual_tests:
        for test in tests:
            if issubclass(tests_dict[test.name].__class__, ManualMessageTest):
                for game in games:
                    run_test_on_game(game, test, tests_dict[test.name])


def run_test_on_game(game: Game, test_entry: Test, test: TestType):
    messages = [Message.from_dict(message_doc) for message_doc in messages_table.search(
        where('game') == game.get_id())]
    if isinstance(test, MessageTest):
        for i, message in enumerate(messages):
            previous_messages = [message.content for message in messages[:i]]
            new_test_result = test.test(message.content, previous_messages)
            test_result_doc = test_results_table.search(
                (where('test') == test_entry.get_id()) & (where('message') == message.get_id()))

            if len(test_result_doc) == 0:
                test_result = TestResult(
                    test_entry.get_id(), message.get_id(), new_test_result)
                test_result.insert_into_table(test_results_table)
            else:
                test_result = TestResult.from_dict(test_result_doc[0])
                test_result.result = new_test_result
                test_result.update_in_table(test_results_table)


if __name__ == "__main__":
    app()
